from __future__ import annotations

from bisect import bisect_left, bisect_right, insort
from collections import deque
import csv
import datetime as dt
import json
import logging
import math
import os
from pathlib import Path
import re
import threading
import time
from typing import Iterable, Optional

from flask import Flask, jsonify, request

try:
    import openpyxl
    from openpyxl.utils.datetime import from_excel
    from openpyxl.utils.cell import get_column_letter
except ImportError as exc:  # pragma: no cover - runtime dependency check
    raise SystemExit(
        "缺少依赖：openpyxl。请先安装 requirements.txt 后再运行。"
    ) from exc

try:
    from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError
except ImportError:  # pragma: no cover - optional dependency for download console
    sync_playwright = None  # type: ignore[assignment]
    PlaywrightTimeoutError = Exception  # type: ignore[assignment]

try:
    from qcloud_cos import CosConfig, CosS3Client
except ImportError:  # pragma: no cover - optional dependency for COS publish
    CosConfig = None  # type: ignore[assignment]
    CosS3Client = None  # type: ignore[assignment]

try:
    from dotenv import load_dotenv
except ImportError:  # pragma: no cover - optional dependency for .env loading
    load_dotenv = None  # type: ignore[assignment]
BASE_DIR = Path(__file__).resolve().parents[1]
INPUT_DIR = BASE_DIR / "input"
OUTPUT_DIR = BASE_DIR / "docs" / "data"
DOCS_DIR = BASE_DIR / "docs"

app = Flask(__name__, static_folder=str(DOCS_DIR), static_url_path="")

OUTPUT_DECIMAL_PLACES = 6

DOWNLOAD_CONFIG_PATH = BASE_DIR / "config" / "download_config.json"
DOWNLOAD_LOCK_PATH = BASE_DIR / "data" / "download.lock"
DOWNLOAD_LOCK_STALE_SECONDS = 60 * 30
DOWNLOAD_DEFAULT_WAIT_MS = 5000
DOWNLOAD_LOGIN_WAIT_SECONDS = 300
DOWNLOAD_LOGIN_STRICT = True
DOWNLOAD_LOGIN_USERNAME = "xuante"

COS_DEFAULT_BUCKET = "anexus-data-1399092305"
COS_DEFAULT_REGION = "ap-guangzhou"
COS_DEFAULT_BASE_PATH = "data"


class DownloadLoginAbort(Exception):
    pass


DOWNLOAD_STATUS = {
    "running": False,
    "run_id": 0,
    "trigger": None,
    "started_at": None,
    "ended_at": None,
    "success": None,
    "message": "",
    "tasks": {},
}
DOWNLOAD_STATUS_LOCK = threading.Lock()


def _cell_to_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, (dt.date, dt.datetime, dt.time)):
        return value.isoformat()
    if isinstance(value, float):
        if math.isnan(value) or math.isinf(value):
            return str(value)
        rounded = round(value, OUTPUT_DECIMAL_PLACES)
        text = f"{rounded:.{OUTPUT_DECIMAL_PLACES}f}"
        return text.rstrip("0").rstrip(".")
    return str(value)

def _round_for_output(value: object) -> object:
    if isinstance(value, float):
        if math.isnan(value) or math.isinf(value):
            return value
        return round(value, OUTPUT_DECIMAL_PLACES)
    return value


def _read_required_env(name: str) -> str:
    value = os.environ.get(name, "").strip()
    if not value:
        raise ValueError(f"缺少环境变量：{name}")
    return value


def _load_cos_settings() -> dict[str, str]:
    if load_dotenv is not None:
        load_dotenv()
    secret_id = _read_required_env("COS_SECRET_ID")
    secret_key = _read_required_env("COS_SECRET_KEY")
    bucket = os.environ.get("COS_BUCKET", COS_DEFAULT_BUCKET).strip() or COS_DEFAULT_BUCKET
    region = os.environ.get("COS_REGION", COS_DEFAULT_REGION).strip() or COS_DEFAULT_REGION
    base_path = os.environ.get("COS_BASE_PATH", COS_DEFAULT_BASE_PATH).strip().strip("/")
    return {
        "secret_id": secret_id,
        "secret_key": secret_key,
        "bucket": bucket,
        "region": region,
        "base_path": base_path,
    }


def _publish_csv_to_cos(local_path: Path, remote_name: str) -> str:
    if CosConfig is None or CosS3Client is None:
        raise ValueError("缺少依赖：cos-python-sdk-v5。请先安装 requirements.txt 后再运行。")
    if not local_path.exists():
        raise FileNotFoundError(f"本地文件不存在：{local_path.name}")

    settings = _load_cos_settings()
    cos_config = CosConfig(
        Region=settings["region"],
        SecretId=settings["secret_id"],
        SecretKey=settings["secret_key"],
        Scheme="https",
    )
    client = CosS3Client(cos_config)
    key = f"{settings['base_path']}/{remote_name}" if settings["base_path"] else remote_name
    with local_path.open("rb") as file_handle:
        client.put_object(
            Bucket=settings["bucket"],
            Body=file_handle,
            Key=key,
            ContentType="text/csv; charset=utf-8",
        )
    return f"https://{settings['bucket']}.cos.{settings['region']}.myqcloud.com/{key}"

def _download_load_config() -> dict:
    if not DOWNLOAD_CONFIG_PATH.exists():
        raise FileNotFoundError("缺少下载配置：config/download_config.json")
    with DOWNLOAD_CONFIG_PATH.open("r", encoding="utf-8") as file_handle:
        return json.load(file_handle)


def _download_save_config(cfg: dict) -> None:
    DOWNLOAD_CONFIG_PATH.parent.mkdir(parents=True, exist_ok=True)
    with DOWNLOAD_CONFIG_PATH.open("w", encoding="utf-8") as file_handle:
        json.dump(cfg, file_handle, ensure_ascii=False, indent=2)


def _download_resolve_path(path_text: str) -> Path:
    path = Path(path_text)
    if path.is_absolute():
        return path
    return BASE_DIR / path


def _download_ensure_dirs(cfg: dict) -> tuple[Path, Path]:
    download_dir = _download_resolve_path(cfg["download_dir"])
    user_data_dir = _download_resolve_path(cfg["user_data_dir"])
    download_dir.mkdir(parents=True, exist_ok=True)
    user_data_dir.mkdir(parents=True, exist_ok=True)
    DOWNLOAD_LOCK_PATH.parent.mkdir(parents=True, exist_ok=True)
    return download_dir, user_data_dir



def _download_acquire_lock() -> bool:
    if DOWNLOAD_LOCK_PATH.exists():
        try:
            content = DOWNLOAD_LOCK_PATH.read_text().strip().split("\n")
            pid = int(content[0]) if content else None
            ts = float(content[1]) if len(content) > 1 else 0
            if pid:
                try:
                    os.kill(pid, 0)
                    return False
                except OSError:
                    # 进程已不存在，直接清锁避免后台任务被卡住
                    DOWNLOAD_LOCK_PATH.unlink(missing_ok=True)
                    pid = None
            if pid is not None:
                if time.time() - ts > DOWNLOAD_LOCK_STALE_SECONDS:
                    DOWNLOAD_LOCK_PATH.unlink(missing_ok=True)
                else:
                    return False
        except Exception:
            try:
                DOWNLOAD_LOCK_PATH.unlink(missing_ok=True)
            except Exception:
                return False
    try:
        fd = os.open(DOWNLOAD_LOCK_PATH, os.O_CREAT | os.O_EXCL | os.O_WRONLY)
        os.write(fd, f"{os.getpid()}\n{time.time()}".encode("utf-8"))
        os.close(fd)
        return True
    except FileExistsError:
        return False


def _download_release_lock() -> None:
    try:
        DOWNLOAD_LOCK_PATH.unlink(missing_ok=True)
    except Exception:
        pass


def _download_set_status(update: dict) -> None:
    with DOWNLOAD_STATUS_LOCK:
        DOWNLOAD_STATUS.update(update)


def _download_update_task_status(task_id: str, update: dict) -> None:
    with DOWNLOAD_STATUS_LOCK:
        if task_id not in DOWNLOAD_STATUS["tasks"]:
            DOWNLOAD_STATUS["tasks"][task_id] = {}
        DOWNLOAD_STATUS["tasks"][task_id].update(update)


def _download_reset_task_status(tasks: list[dict]) -> None:
    with DOWNLOAD_STATUS_LOCK:
        DOWNLOAD_STATUS["tasks"] = {
            task["id"]: {
                "status": "idle",
                "message": "",
                "file": "",
                "validation": None,
                "started_at": None,
                "ended_at": None,
            }
            for task in tasks
        }


def _download_now_ts() -> str:
    return dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def _download_click_button_by_text(page, text: str) -> None:
    try:
        loc = page.get_by_role("button", name=text)
        if loc.count() > 0:
            loc.first.click(timeout=5000)
            return
    except Exception:
        pass
    page.get_by_text(text, exact=True).first.click(timeout=5000)


def _download_click_by_xpath(page, xpath: str) -> bool:
    if not xpath:
        return False
    selectors = [f"xpath={xpath}", xpath]
    for selector in selectors:
        try:
            loc = page.locator(selector)
            if loc.count() > 0:
                target = loc.first
                try:
                    target.scroll_into_view_if_needed(timeout=2000)
                except Exception:
                    pass
                target.click(timeout=5000, force=True)
                return True
        except Exception:
            pass
    return False


def _download_click_download_excel(page) -> None:
    try:
        loc = page.get_by_role("button", name=re.compile("EXCEL"))
        if loc.count() > 0:
            loc.first.click(timeout=5000)
            return
    except Exception:
        pass
    page.get_by_text(re.compile("EXCEL"), exact=False).first.click(timeout=5000)


def _download_login_required(page) -> bool:
    keywords = ["微信扫码", "微信扫一扫", "扫码登录", "二维码有效期", "QQ登录", "用户密码登录"]
    if page.is_closed():
        return False
    if _download_is_logged_in(page):
        return False
    for keyword in keywords:
        try:
            if page.get_by_text(keyword, exact=False).is_visible():
                return True
        except Exception:
            pass
    if DOWNLOAD_LOGIN_STRICT:
        try:
            if page.get_by_role("button", name="登录").is_visible():
                return True
        except Exception:
            pass
    return False


def _download_is_logged_in(page) -> bool:
    if page.is_closed():
        return False
    if not DOWNLOAD_LOGIN_USERNAME:
        return False
    try:
        if page.get_by_text(DOWNLOAD_LOGIN_USERNAME, exact=False).is_visible():
            return True
    except Exception:
        pass
    return False


def _download_wait_for_login(page) -> None:
    start = time.time()
    while True:
        if page.is_closed():
            raise DownloadLoginAbort("登录窗口被关闭，已中止本次任务")
        if not _download_login_required(page):
            return
        if time.time() - start > DOWNLOAD_LOGIN_WAIT_SECONDS:
            raise DownloadLoginAbort("登录超时，请重新扫码登录")
        time.sleep(2)


def _download_ensure_logged_in(context, url: str) -> None:
    page = context.new_page()
    page.bring_to_front()
    page.goto(url, wait_until="domcontentloaded")
    try:
        page.wait_for_load_state("networkidle", timeout=DOWNLOAD_DEFAULT_WAIT_MS)
    except PlaywrightTimeoutError:
        pass
    if _download_login_required(page):
        _download_set_status({"message": "需要登录：请在弹出的浏览器中扫码登录后再继续。"})
        _download_wait_for_login(page)
    page.close()


def _download_wait_chart_update(page) -> None:
    try:
        page.wait_for_load_state("networkidle", timeout=DOWNLOAD_DEFAULT_WAIT_MS)
    except PlaywrightTimeoutError:
        pass
    page.wait_for_timeout(DOWNLOAD_DEFAULT_WAIT_MS)


def _download_open_freq_dropdown(page) -> None:
    candidates = ["周", "月", "日"]
    try:
        loc = page.locator("input.el-input__inner[placeholder='粒度']")
        if loc.count() > 0:
            loc.first.click(timeout=5000)
            return
    except Exception:
        pass

    try:
        loc = page.locator("div.wa-chart-toolbox").locator(
            "input.el-input__inner[placeholder='粒度']"
        )
        if loc.count() > 0:
            loc.first.click(timeout=5000)
            return
    except Exception:
        pass

    for text in candidates:
        try:
            loc = page.get_by_role("button", name=text)
            if loc.count() > 0:
                loc.first.click(timeout=5000)
                return
        except Exception:
            pass
    for text in candidates:
        try:
            loc = page.get_by_text(text, exact=True)
            if loc.count() > 0:
                loc.first.click(timeout=5000)
                return
        except Exception:
            pass

    for text in candidates:
        try:
            loc = page.locator("[aria-haspopup], [class*='select'], [class*='dropdown'], [class*='picker']").filter(
                has_text=re.compile(f"^{text}$")
            )
            if loc.count() > 0:
                loc.first.click(timeout=5000)
                return
        except Exception:
            pass

    try:
        export_img = page.get_by_text("导出图片", exact=False)
        if export_img.count() > 0:
            loc = export_img.first.locator(
                "xpath=preceding::*[(self::div or self::span or self::button) "
                "and (contains(normalize-space(.), '周') or contains(normalize-space(.), '月') or contains(normalize-space(.), '日'))][1]"
            )
            if loc.count() > 0:
                loc.first.click(timeout=5000)
                return
    except Exception:
        pass

    try:
        export_excel = page.get_by_text(re.compile("EXCEL"), exact=False)
        if export_excel.count() > 0:
            loc = export_excel.first.locator(
                "xpath=preceding::*[(self::div or self::span or self::button) "
                "and (contains(normalize-space(.), '周') or contains(normalize-space(.), '月') or contains(normalize-space(.), '日'))][1]"
            )
            if loc.count() > 0:
                loc.first.click(timeout=5000)
                return
    except Exception:
        pass

    try:
        css = "body > div.wa-page > div.wa-container.wa-content > div:nth-child(2) > div > div.el-row > div > div:nth-child(1) > div.wa-chart-toolbox.text-right.mr20 > form > div > div > div.el-badge.item > div > div > input"
        loc = page.locator(css)
        if loc.count() > 0:
            loc.first.click(timeout=5000)
            return
    except Exception:
        pass

    raise RuntimeError("未找到频率下拉菜单入口，请确认页面上“周/月/日”控件可见")


def _download_select_dropdown_option(page, option_text: str) -> None:
    try:
        loc = page.locator("li.el-select-dropdown__item").filter(
            has_text=re.compile(f"^{option_text}$")
        )
        if loc.count() > 0:
            loc.first.click(timeout=5000)
            return
    except Exception:
        pass

    try:
        loc = page.get_by_role("option", name=option_text)
        if loc.count() > 0:
            loc.first.click(timeout=5000)
            return
    except Exception:
        pass

    try:
        loc = page.locator("li").filter(has_text=re.compile(f"^{option_text}$"))
        if loc.count() > 0:
            loc.first.click(timeout=5000)
            return
    except Exception:
        pass
    page.get_by_text(option_text, exact=True).last.click(timeout=5000)


def _download_parse_date_str(text: str) -> Optional[dt.date]:
    text = text.strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y%m%d"):
        try:
            return dt.datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _download_read_date_column(path: Path) -> list[dt.date]:
    workbook = openpyxl.load_workbook(filename=path, read_only=True, data_only=True)
    sheet = workbook.active
    dates: list[dt.date] = []
    for index, row in enumerate(sheet.iter_rows(min_row=1, max_col=1, values_only=True), start=1):
        value = row[0]
        if index == 1 and isinstance(value, str):
            continue
        if value is None:
            continue
        if isinstance(value, dt.datetime):
            dates.append(value.date())
        elif hasattr(value, "year") and hasattr(value, "month") and hasattr(value, "day"):
            dates.append(dt.date(value.year, value.month, value.day))
        elif isinstance(value, str):
            parsed = _download_parse_date_str(value)
            if parsed:
                dates.append(parsed)
    workbook.close()
    return sorted(set(dates))


def _download_validate_dates(dates: list[dt.date], rule: dict) -> dict:
    result = {"ok": True, "errors": [], "warnings": [], "stats": {}}
    if not dates:
        return {"ok": False, "errors": ["日期列为空"], "warnings": [], "stats": {}}

    min_date, max_date = dates[0], dates[-1]
    span_days = (max_date - min_date).days
    result["stats"].update({"min": str(min_date), "max": str(max_date), "span_days": span_days})

    min_span_years = rule.get("min_span_years")
    if min_span_years is not None:
        min_days = min_span_years * 365.25
        if span_days < min_days:
            result["ok"] = False
            result["errors"].append(f"时间跨度不足{min_span_years}年")

    min_start_date = rule.get("min_start_date")
    if min_start_date:
        try:
            target = dt.datetime.strptime(min_start_date, "%Y-%m-%d").date()
            if min_date > target:
                result["ok"] = False
                result["errors"].append(f"起始日期晚于{min_start_date}")
        except ValueError:
            result["warnings"].append("起始日期规则格式无效")

    expected_freq = rule.get("freq")
    ratio_threshold = rule.get("freq_ratio", 0.8)
    if expected_freq:
        diffs = []
        for i in range(1, len(dates)):
            diff = (dates[i] - dates[i - 1]).days
            if diff > 0:
                diffs.append(diff)
        if not diffs:
            return {"ok": False, "errors": ["无法计算频率"], "warnings": [], "stats": result["stats"]}

        if expected_freq == "day":
            match = sum(1 for diff in diffs if 1 <= diff <= 4)
            ratio = match / len(diffs)
            result["stats"].update({"daily_ratio": round(ratio, 4)})
            if ratio < ratio_threshold:
                result["ok"] = False
                result["errors"].append("日频校验未通过（1~4天占比过低）")
        elif expected_freq == "week":
            match = sum(1 for diff in diffs if 5 <= diff <= 9)
            ratio = match / len(diffs)
            result["stats"].update({"weekly_ratio": round(ratio, 4)})
            if ratio < ratio_threshold:
                result["ok"] = False
                result["errors"].append("周频校验未通过（5~9天占比过低）")

    return result


def _download_perform_task(task: dict, url: str, context, download_dir: Path) -> dict:
    task_id = task["id"]
    page = context.new_page()
    page.bring_to_front()
    page.goto(url, wait_until="domcontentloaded")
    try:
        page.wait_for_load_state("networkidle", timeout=DOWNLOAD_DEFAULT_WAIT_MS)
    except PlaywrightTimeoutError:
        pass
    if _download_login_required(page):
        _download_update_task_status(task_id, {"status": "waiting_login", "message": "等待扫码登录"})
        _download_set_status({"message": "需要登录：请在弹出的浏览器中扫码登录后再继续。"})
        _download_wait_for_login(page)


    if task_id == "gdp":
        pass
    elif task_id == "market_amount":
        _download_click_button_by_text(page, "所有")
        _download_wait_chart_update(page)
    elif task_id == "margin_trading":
        _download_click_button_by_text(page, "20年")
        _download_wait_chart_update(page)
    elif task_id == "all_a_index":
        _download_click_button_by_text(page, "所有")
        _download_wait_chart_update(page)
        _download_open_freq_dropdown(page)
        _download_select_dropdown_option(page, "日")
        _download_wait_chart_update(page)
    elif task_id == "national_debt":
        _download_click_button_by_text(page, "20年")
        _download_wait_chart_update(page)
    elif task_id in ("sp500_index", "nasdaq_index"):
        if not _download_click_by_xpath(page, str(task.get("range_xpath", "")).strip()):
            _download_click_button_by_text(page, "20年")
        _download_wait_chart_update(page)
        _download_open_freq_dropdown(page)
        _download_select_dropdown_option(page, "日")
        _download_wait_chart_update(page)
    else:
        raise RuntimeError(f"未知任务: {task_id}")

    try:
        with page.expect_download(timeout=30000) as dl_info:
            _download_click_download_excel(page)
    except PlaywrightTimeoutError:
        if _download_login_required(page):
            raise RuntimeError("需要登录：请扫码登录后重试")
        raise RuntimeError("下载超时，请确认页面已刷新并重试")
    download = dl_info.value
    suggested = download.suggested_filename or f"{task_id}.xlsx"
    ts = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{task_id}_{ts}_{suggested}"
    dest = download_dir / filename
    download.save_as(dest)

    dates = _download_read_date_column(dest)
    validation = _download_validate_dates(dates, task.get("validation", {}))

    final_path = dest
    output_name = task.get("output_name")
    if validation.get("ok") and output_name:
        suffix = dest.suffix or ".xlsx"
        final_path = dest.parent / f"{output_name}{suffix}"
        try:
            final_path.unlink(missing_ok=True)
        except Exception:
            pass
        dest.replace(final_path)

    page.close()
    return {"file": str(final_path), "validation": validation}


class _DownloadRunner:
    def __init__(self) -> None:
        self.thread: Optional[threading.Thread] = None

    def start(self, task_ids: list[str], urls: dict[str, str], trigger: str) -> bool:
        with DOWNLOAD_STATUS_LOCK:
            if DOWNLOAD_STATUS["running"]:
                return False
            DOWNLOAD_STATUS["running"] = True
            DOWNLOAD_STATUS["run_id"] += 1
            DOWNLOAD_STATUS["trigger"] = trigger
            DOWNLOAD_STATUS["started_at"] = _download_now_ts()
            DOWNLOAD_STATUS["ended_at"] = None
            DOWNLOAD_STATUS["success"] = None
            DOWNLOAD_STATUS["message"] = ""
        self.thread = threading.Thread(
            target=self._run, args=(task_ids, urls, trigger, DOWNLOAD_STATUS["run_id"]), daemon=True
        )
        self.thread.start()
        return True

    def _run(self, task_ids: list[str], urls: dict[str, str], trigger: str, run_id: int) -> None:
        cfg = _download_load_config()
        download_dir, user_data_dir = _download_ensure_dirs(cfg)
        tasks = {task["id"]: task for task in cfg["tasks"]}
        _download_reset_task_status(cfg["tasks"])

        if sync_playwright is None:
            _download_set_status({
                "running": False,
                "ended_at": _download_now_ts(),
                "success": False,
                "message": "缺少依赖：playwright。请先执行 pip install -r requirements.txt",
            })
            return

        if not _download_acquire_lock():
            _download_set_status({
                "running": False,
                "ended_at": _download_now_ts(),
                "success": False,
                "message": "已有任务在运行或残留锁文件，请稍后再试或清理 data/download.lock",
            })
            return

        overall_success = True
        try:
            with sync_playwright() as p:
                context = p.chromium.launch_persistent_context(
                    user_data_dir=str(user_data_dir),
                    headless=cfg.get("headless", False),
                    accept_downloads=True,
                    channel="chrome",
                )
                first_url = cfg["tasks"][0]["url"] if cfg.get("tasks") else None
                if first_url:
                    _download_ensure_logged_in(context, first_url)
                    _download_set_status({"message": ""})
                abort_all = False
                for task_id in task_ids:
                    task = tasks.get(task_id)
                    if not task:
                        continue
                    task_url = urls.get(task_id) or task["url"]
                    _download_update_task_status(task_id, {"status": "running", "started_at": _download_now_ts()})
                    try:
                        result = _download_perform_task(task, task_url, context, download_dir)
                        validation = result["validation"]
                        if not validation.get("ok", True):
                            overall_success = False
                            err_msg = "; ".join(validation.get("errors", [])) or "校验失败"
                            _download_update_task_status(task_id, {
                                "status": "error",
                                "ended_at": _download_now_ts(),
                                "message": f"校验失败：{err_msg}",
                                "file": result["file"],
                                "validation": validation,
                            })
                        else:
                            _download_update_task_status(task_id, {
                                "status": "success",
                                "ended_at": _download_now_ts(),
                                "file": result["file"],
                                "validation": validation,
                            })
                    except DownloadLoginAbort as exc:
                        overall_success = False
                        _download_update_task_status(task_id, {
                            "status": "error",
                            "ended_at": _download_now_ts(),
                            "message": str(exc),
                        })
                        _download_set_status({"message": str(exc)})
                        abort_all = True
                    except Exception as exc:
                        overall_success = False
                        _download_update_task_status(task_id, {
                            "status": "error",
                            "ended_at": _download_now_ts(),
                            "message": str(exc),
                        })
                    if abort_all:
                        break
                context.close()
        except Exception as exc:
            overall_success = False
            _download_set_status({"message": f"运行失败: {exc}"})
        finally:
            _download_release_lock()

        _download_set_status({
            "running": False,
            "ended_at": _download_now_ts(),
            "success": overall_success,
            "message": "完成" if overall_success else "部分任务失败",
        })


download_runner = _DownloadRunner()


def _is_garbled_text(text: str) -> bool:
    if "\ufffd" in text:
        return True
    for char in text:
        code_point = ord(char)
        if code_point < 32 and char not in ("\t", "\n", "\r"):
            return True
    return False


def _parse_date(value: object, *, epoch: dt.datetime) -> dt.date:
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date) and not isinstance(value, dt.datetime):
        return value
    if isinstance(value, bool):
        raise ValueError("布尔类型不是有效日期")
    if isinstance(value, (int, float)):
        if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
            raise ValueError("数值为 NaN/Inf")
        return from_excel(value, epoch=epoch).date()
    if isinstance(value, str):
        text = value.strip()
        if not text:
            raise ValueError("日期为空白")
        candidates = (
            "%Y-%m-%d",
            "%Y/%m/%d",
            "%Y.%m.%d",
            "%Y%m%d",
            "%Y-%m-%d %H:%M:%S",
            "%Y/%m/%d %H:%M:%S",
        )
        for fmt in candidates:
            try:
                parsed = dt.datetime.strptime(text, fmt)
                return parsed.date()
            except ValueError:
                continue
        try:
            return dt.date.fromisoformat(text)
        except ValueError as exc:
            raise ValueError(f"无法解析日期：{text}") from exc
    raise ValueError(f"不支持的日期类型：{type(value).__name__}")


def _validate_text_or_number(value: object) -> object:
    if value is None:
        raise ValueError("内容空白")
    if isinstance(value, bool):
        raise ValueError("不支持布尔类型")
    if isinstance(value, str):
        text = value.strip()
        if not text:
            raise ValueError("内容空白")
        if _is_garbled_text(text):
            raise ValueError("疑似乱码/控制字符")
        return text
    if isinstance(value, (int, float)):
        if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
            raise ValueError("数值为 NaN/Inf")
        return value
    raise ValueError(f"不支持的类型：{type(value).__name__}")


def _validate_header_cell(value: object) -> str:
    if value is None:
        raise ValueError("标题空白")
    if not isinstance(value, str):
        raise ValueError("标题必须为文本")
    text = value.strip()
    if not text:
        raise ValueError("标题空白")
    if _is_garbled_text(text):
        raise ValueError("标题疑似乱码/控制字符")
    return text


def process_xlsx_to_outputs(source_path: Path, output_csv_path: Path, output_xlsx_path: Path) -> None:
    workbook = openpyxl.load_workbook(source_path, data_only=True, read_only=True)
    sheet_name = workbook.sheetnames[0] if workbook.sheetnames else None
    if not sheet_name:
        raise ValueError("未找到可用工作表")

    sheet = workbook[sheet_name]
    rows_iter = sheet.iter_rows(values_only=True)
    header_values = next(rows_iter, None)
    if not header_values:
        raise ValueError("未找到标题行")

    def _is_blank(value: object) -> bool:
        return value is None or (isinstance(value, str) and not value.strip())

    last_col = 0
    for column_index, value in enumerate(header_values, start=1):
        if not _is_blank(value):
            last_col = column_index

    if last_col == 0:
        raise ValueError("标题行为空")

    columns_to_keep: list[int] = [col for col in range(1, last_col + 1) if col not in (2, 3, 4)]
    if 1 not in columns_to_keep:
        columns_to_keep.insert(0, 1)

    output_rows: list[list[object]] = []
    row_dates: list[dt.date] = []

    kept_header: list[str] = []
    for col in columns_to_keep:
        value = header_values[col - 1] if col - 1 < len(header_values) else None
        coordinate = f"{get_column_letter(col)}1"
        try:
            kept_header.append(_validate_header_cell(value))
        except ValueError as exc:
            raise ValueError(f"{coordinate} 标题错误：{exc}") from exc

    output_rows.append(kept_header)

    for row_offset, row_values in enumerate(rows_iter, start=2):
        values = list(row_values[:last_col])
        if len(values) < last_col:
            values.extend([None] * (last_col - len(values)))

        kept_values = [values[col - 1] for col in columns_to_keep]
        if all(_is_blank(value) for value in kept_values):
            continue

        normalized_row: list[object] = []
        for position, col in enumerate(columns_to_keep):
            value = values[col - 1]
            coordinate = f"{get_column_letter(col)}{row_offset}"
            try:
                if position == 0:
                    parsed = _parse_date(value, epoch=workbook.epoch)
                    normalized_row.append(parsed.isoformat())
                else:
                    normalized_row.append(_validate_text_or_number(value))
            except ValueError as exc:
                raise ValueError(f"{coordinate} 内容错误：{exc}") from exc

        output_rows.append(normalized_row)
        row_dates.append(dt.date.fromisoformat(normalized_row[0]))

    if len(output_rows) <= 1:
        raise ValueError("没有可导出的数据行")

    data_rows = output_rows[1:]
    if len(data_rows) != len(row_dates):
        raise ValueError("内部错误：行数不一致")

    sorted_data_rows = [row for _, row in sorted(zip(row_dates, data_rows), key=lambda item: item[0])]
    final_rows = [output_rows[0], *sorted_data_rows]

    output_csv_path.parent.mkdir(parents=True, exist_ok=True)
    with output_csv_path.open("w", encoding="utf-8-sig", newline="") as file_handle:
        writer = csv.writer(file_handle)
        for row in final_rows:
            writer.writerow([_cell_to_text(value) for value in row])

    workbook_out = openpyxl.Workbook()
    sheet_out = workbook_out.active
    sheet_out.title = "processed"
    sheet_out.freeze_panes = "B2"
    for row in final_rows:
        sheet_out.append([_round_for_output(value) for value in row])
    output_xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    workbook_out.save(output_xlsx_path)

def _find_input_xlsx(stem: str) -> Path:
    if not INPUT_DIR.exists():
        raise FileNotFoundError("input/ 目录不存在")

    def normalize(text: str) -> str:
        return " ".join(text.strip().split()).lower()

    candidates: list[Path] = []
    for path in INPUT_DIR.iterdir():
        if not path.is_file():
            continue
        if path.name.startswith("~$"):
            continue
        if path.suffix.lower() != ".xlsx":
            continue
        if normalize(path.stem) == normalize(stem):
            candidates.append(path)

    if not candidates:
        raise FileNotFoundError(f"未找到文件：{stem}.xlsx（请放入 input/）")
    if len(candidates) > 1:
        raise FileNotFoundError(f"找到多个匹配文件：{stem}.xlsx（请保留一个）")
    return candidates[0]


def _coerce_float(value: object) -> float:
    if isinstance(value, bool):
        raise ValueError("不支持布尔类型")
    if value is None:
        raise ValueError("内容空白")
    if isinstance(value, (int, float)):
        if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
            raise ValueError("数值为 NaN/Inf")
        return float(value)
    if isinstance(value, str):
        text = value.strip()
        if not text:
            raise ValueError("内容空白")
        if _is_garbled_text(text):
            raise ValueError("疑似乱码/控制字符")
        cleaned = text.replace(",", "")
        if cleaned.endswith("%"):
            cleaned = cleaned[:-1].strip()
        try:
            return float(cleaned)
        except ValueError as exc:
            raise ValueError(f"无法解析为数值：{text}") from exc
    raise ValueError(f"不支持的类型：{type(value).__name__}")


def _normalize_yield(yield_raw: float) -> float:
    if yield_raw > 1.0:
        return yield_raw / 100.0
    return yield_raw


def _iter_rows_values(sheet: object, *, last_col: int) -> Iterable[tuple[object, ...]]:
    for row_values in sheet.iter_rows(values_only=True):
        values = tuple(row_values[:last_col])
        if len(values) < last_col:
            values = values + (None,) * (last_col - len(values))
        yield values


def _validate_expected_header(actual: object, expected: str, coordinate: str) -> None:
    text = _validate_header_cell(actual)
    if text != expected:
        raise ValueError(f"{coordinate} 标题不匹配：期望“{expected}”，实际“{text}”")


def _rolling_percentile(sorted_window: list[float], value: float) -> float:
    window_size = len(sorted_window)
    if window_size <= 0:
        raise ValueError("窗口为空")
    if window_size == 1:
        return 50.0
    left = bisect_left(sorted_window, value)
    right = bisect_right(sorted_window, value)
    rank_low = left + 1
    rank_high = right
    avg_rank = (rank_low + rank_high) / 2.0
    return 100.0 * (avg_rank - 1.0) / (window_size - 1.0)


def _moving_average(values: list[float], window: int) -> list[float | None]:
    if window <= 0:
        raise ValueError("移动平均窗口必须为正整数")
    out: list[float | None] = []
    q: deque[float] = deque()
    sum_values = 0.0
    for value in values:
        q.append(value)
        sum_values += value
        if len(q) > window:
            sum_values -= q.popleft()
        if len(q) == window:
            out.append(sum_values / window)
        else:
            out.append(None)
    return out


def _rolling_percentiles(values: list[float | None], window: int) -> list[float | None]:
    if window <= 0:
        raise ValueError("滚动窗口必须为正整数")

    first_valid = 0
    while first_valid < len(values) and values[first_valid] is None:
        first_valid += 1

    out: list[float | None] = [None] * len(values)
    if first_valid >= len(values):
        return out

    sorted_window: list[float] = []
    q: deque[float] = deque()

    for index in range(first_valid, len(values)):
        current = values[index]
        if current is None:
            out[index] = None
            continue

        insort(sorted_window, float(current))
        q.append(float(current))
        if len(q) > window:
            leaving = q.popleft()
            remove_index = bisect_left(sorted_window, leaving)
            if remove_index >= len(sorted_window) or sorted_window[remove_index] != leaving:
                raise ValueError("内部错误：滚动窗口移除失败")
            sorted_window.pop(remove_index)

        if len(q) < window:
            out[index] = None
            continue

        out[index] = _rolling_percentile(sorted_window, float(current))

    return out


def _rolling_percentiles_with_min_window(
    values: list[float | None],
    window: int,
    min_window: int,
) -> list[float | None]:
    if window <= 0:
        raise ValueError("滚动窗口必须为正整数")
    if min_window <= 0:
        raise ValueError("最小递减滚动周期必须为正整数")
    if min_window > window:
        raise ValueError("最小递减滚动周期不能大于分位滚动周期")

    first_valid = 0
    while first_valid < len(values) and values[first_valid] is None:
        first_valid += 1

    out: list[float | None] = [None] * len(values)
    if first_valid >= len(values):
        return out

    sorted_window: list[float] = []
    q: deque[float] = deque()

    for index in range(first_valid, len(values)):
        current = values[index]
        if current is None:
            out[index] = None
            continue

        insort(sorted_window, float(current))
        q.append(float(current))
        if len(q) > window:
            leaving = q.popleft()
            remove_index = bisect_left(sorted_window, leaving)
            if remove_index >= len(sorted_window) or sorted_window[remove_index] != leaving:
                raise ValueError("内部错误：滚动窗口移除失败")
            sorted_window.pop(remove_index)

        if len(q) < min_window:
            out[index] = None
            continue

        out[index] = _rolling_percentile(sorted_window, float(current))

    return out


def _rolling_percentiles_with_min_window_sizes(
    values: list[float | None],
    window: int,
    min_window: int,
) -> tuple[list[float | None], list[int | None]]:
    if window <= 0:
        raise ValueError("滚动窗口必须为正整数")
    if min_window <= 0:
        raise ValueError("最小递减滚动周期必须为正整数")
    if min_window > window:
        raise ValueError("最小递减滚动周期不能大于分位滚动周期")

    first_valid = 0
    while first_valid < len(values) and values[first_valid] is None:
        first_valid += 1

    out: list[float | None] = [None] * len(values)
    window_sizes: list[int | None] = [None] * len(values)
    if first_valid >= len(values):
        return out, window_sizes

    sorted_window: list[float] = []
    q: deque[float] = deque()

    for index in range(first_valid, len(values)):
        current = values[index]
        if current is None:
            out[index] = None
            continue

        insort(sorted_window, float(current))
        q.append(float(current))
        if len(q) > window:
            leaving = q.popleft()
            remove_index = bisect_left(sorted_window, leaving)
            if remove_index >= len(sorted_window) or sorted_window[remove_index] != leaving:
                raise ValueError("内部错误：滚动窗口移除失败")
            sorted_window.pop(remove_index)

        if len(q) < min_window:
            out[index] = None
            continue

        out[index] = _rolling_percentile(sorted_window, float(current))
        window_sizes[index] = len(q)

    return out, window_sizes



def _load_ratio_series(source_path: Path) -> tuple[list[str], list[float]]:
    workbook = openpyxl.load_workbook(source_path, data_only=True, read_only=True)
    try:
        sheet_name = workbook.sheetnames[0] if workbook.sheetnames else None
        if not sheet_name:
            raise ValueError(f"{source_path.name}：未找到可用工作表")
        sheet = workbook[sheet_name]
        epoch = workbook.epoch

        last_col = 4  # A-D
        rows_iter = _iter_rows_values(sheet, last_col=last_col)
        header = next(rows_iter, None)
        if not header:
            raise ValueError(f"{source_path.name}：未找到标题行")

        header_a = _validate_header_cell(header[0])
        _ = _validate_header_cell(header[3])
        if "日期" not in header_a and header_a.lower() != "date":
            raise ValueError(f"{source_path.name}：A1 标题应为“日期”")

        rows: list[tuple[dt.date, float]] = []
        for _, values in enumerate(rows_iter, start=2):
            try:
                date = _parse_date(values[0], epoch=epoch)
                ratio = _coerce_float(values[3])
                rows.append((date, ratio))
            except Exception:
                continue

        if not rows:
            raise ValueError(f"{source_path.name}：清洗后没有可用数据行")

        rows.sort(key=lambda item: item[0])
        dates = [date.isoformat() for date, _ in rows]
        metrics = [metric for _, metric in rows]
        return dates, metrics
    finally:
        workbook.close()


def _load_erp_series() -> tuple[list[str], list[float], list[float], list[float], list[float]]:
    pe_path = _find_input_xlsx("data_PE")
    bond_path = _find_input_xlsx("data_bond")

    pe_rows = _process_data_pe(pe_path)
    bond_rows = _process_data_bond(bond_path)
    merged_rows = _merge_by_bond_dates(bond_rows, pe_rows)
    erp_rows = _compute_erp_rows(merged_rows, bond_rows)

    dates: list[str] = []
    erp_values: list[float] = []
    bond_yield_values: list[float] = []
    pe_values: list[float] = []
    close_values: list[float] = []
    for row_index, row in enumerate(erp_rows[1:], start=2):
        try:
            date_text = str(row[0])
            _ = dt.date.fromisoformat(date_text)
            value = row[4]
            if not isinstance(value, (int, float)):
                raise ValueError("数值类型不合法")
            dates.append(date_text)
            erp_values.append(float(value))

            yield_value = row[1]
            if not isinstance(yield_value, (int, float)):
                raise ValueError("十年期收益率类型不合法")
            bond_yield_values.append(float(yield_value))

            pe_value = row[2]
            if not isinstance(pe_value, (int, float)):
                raise ValueError("PE 类型不合法")
            pe_values.append(float(pe_value))

            close_value = row[3]
            if not isinstance(close_value, (int, float)):
                raise ValueError("收盘点位类型不合法")
            close_values.append(float(close_value))
        except Exception as exc:
            raise ValueError(f"ERP 第 {row_index} 行数据不合法：{exc}") from exc

    if not dates:
        raise ValueError("ERP 数据为空")
    return dates, erp_values, bond_yield_values, pe_values, close_values


def _nearest_index(dates: list[dt.date], target: dt.date) -> int:
    if not dates:
        raise ValueError("日期序列为空")
    index = bisect_left(dates, target)
    if index <= 0:
        return 0
    if index >= len(dates):
        return len(dates) - 1
    before = dates[index - 1]
    after = dates[index]
    diff_before = abs((target - before).days)
    diff_after = abs((after - target).days)
    if diff_before <= diff_after:
        return index - 1
    return index


def _build_percentile_records(
    dates: list[str],
    values: list[float],
    *,
    ma_window: int,
    rp_window: int,
    internal_mode: str,
    min_window: int | None,
) -> list[tuple[dt.date, float]]:
    ma_values = _moving_average(values, ma_window)
    if internal_mode == "auto":
        pct_values = _rolling_percentiles(ma_values, rp_window)
    else:
        if min_window is None:
            raise ValueError("最小递减滚动周期不能为空")
        pct_values = _rolling_percentiles_with_min_window(ma_values, rp_window, min_window)
    out: list[tuple[dt.date, float]] = []
    for index, date_text in enumerate(dates):
        pct = pct_values[index]
        if pct is None:
            continue
        out.append((dt.date.fromisoformat(date_text), float(pct)))
    return out


def _keep_weekly_latest_records(records: list[tuple[dt.date, float]]) -> list[tuple[dt.date, float]]:
    if not records:
        return records
    latest_by_week: dict[tuple[int, int], tuple[dt.date, float]] = {}
    for date_value, metric_value in records:
        iso_year, iso_week, _ = date_value.isocalendar()
        latest_by_week[(iso_year, iso_week)] = (date_value, metric_value)
    return sorted(latest_by_week.values(), key=lambda item: item[0])


def _keep_weekly_latest_rows(rows: list[list[object]]) -> list[list[object]]:
    if len(rows) <= 2:
        return rows
    header = rows[0]
    latest_by_week: dict[tuple[int, int], list[object]] = {}
    for row in rows[1:]:
        if not row:
            continue
        try:
            date_value = dt.date.fromisoformat(str(row[0]))
        except ValueError:
            continue
        iso_year, iso_week, _ = date_value.isocalendar()
        latest_by_week[(iso_year, iso_week)] = row
    out = [header]
    out.extend(sorted(latest_by_week.values(), key=lambda row: dt.date.fromisoformat(str(row[0]))))
    return out


def _build_erp_percentile_records(
    dates: list[str],
    erp_values: list[float],
    yields: list[float],
    closes: list[float],
    *,
    ma_window: int,
    rp_window: int,
    internal_mode: str,
    min_window: int | None,
) -> list[dict[str, object]]:
    ma_values = _moving_average(erp_values, ma_window)
    if internal_mode == "auto":
        pct_values = _rolling_percentiles(ma_values, rp_window)
    else:
        if min_window is None:
            raise ValueError("最小递减滚动周期不能为空")
        pct_values = _rolling_percentiles_with_min_window(ma_values, rp_window, min_window)
    out: list[dict[str, object]] = []
    for index, date_text in enumerate(dates):
        pct = pct_values[index]
        if pct is None:
            continue
        out.append(
            {
                "date": dt.date.fromisoformat(date_text),
                "erp_percentile": float(pct),
                "erp": float(erp_values[index]),
                "yield": float(yields[index]),
                "close": float(closes[index]),
            }
        )
    return out

def _process_data_pe(source_path: Path) -> list[tuple[dt.date, float, float]]:
    workbook = openpyxl.load_workbook(source_path, data_only=True, read_only=True)
    sheet_name = workbook.sheetnames[0] if workbook.sheetnames else None
    if not sheet_name:
        raise ValueError("data_PE：未找到可用工作表")
    sheet = workbook[sheet_name]

    last_col = 8
    rows_iter = _iter_rows_values(sheet, last_col=last_col)
    header = next(rows_iter, None)
    if not header:
        raise ValueError("data_PE：未找到标题行")

    _validate_expected_header(header[0], "日期", "A1")
    _validate_expected_header(header[3], "PE-TTM-S", "D1")
    _validate_expected_header(header[7], "收盘点位", "H1")

    fill_dates = [
        "2018-08-03",
        "2018-08-06",
        "2018-08-07",
        "2018-08-08",
        "2018-08-09",
        "2018-08-10",
        "2018-08-13",
        "2018-08-14",
        "2018-08-15",
        "2018-08-16",
        "2018-08-17",
        "2018-08-20",
        "2018-08-21",
        "2018-08-22",
        "2018-08-23",
        "2018-08-24",
    ]
    fill_values = [
        3892.88,
        3828.14,
        3933.12,
        3871.35,
        3963.8,
        3979.61,
        3978.56,
        3962.88,
        3876.46,
        3846.75,
        3785.01,
        3814.7,
        3870.75,
        3838.79,
        3856.65,
        3854.99,
    ]
    fill_close_by_date = {
        dt.date.fromisoformat(date): value for date, value in zip(fill_dates, fill_values)
    }

    rows: list[tuple[dt.date, float, float]] = []
    for row_index, values in enumerate(rows_iter, start=2):
        if all(value is None or (isinstance(value, str) and not value.strip()) for value in values):
            continue

        date = _parse_date(values[0], epoch=workbook.epoch)
        try:
            pe = _coerce_float(values[3])
        except ValueError as exc:
            raise ValueError(f"data_PE D{row_index} 内容错误：{exc}") from exc

        close_value = values[7]
        if date in fill_close_by_date and (
            close_value is None or (isinstance(close_value, str) and not close_value.strip())
        ):
            close_value = fill_close_by_date[date]
        try:
            close = _coerce_float(close_value)
        except ValueError as exc:
            raise ValueError(f"data_PE H{row_index} 内容错误：{exc}") from exc

        if pe <= 0:
            raise ValueError(f"data_PE D{row_index} 内容错误：PE 必须为正数")

        rows.append((date, pe, close))

    if not rows:
        raise ValueError("data_PE：没有可用数据行")

    rows.sort(key=lambda item: item[0])
    return rows


def _process_data_bond(source_path: Path) -> list[tuple[dt.date, float, float]]:
    workbook = openpyxl.load_workbook(source_path, data_only=True, read_only=True)
    sheet_name = workbook.sheetnames[0] if workbook.sheetnames else None
    if not sheet_name:
        raise ValueError("data_bond：未找到可用工作表")
    sheet = workbook[sheet_name]

    last_col = 5
    rows_iter = _iter_rows_values(sheet, last_col=last_col)
    header = next(rows_iter, None)
    if not header:
        raise ValueError("data_bond：未找到标题行")

    _validate_expected_header(header[0], "日期", "A1")
    _validate_expected_header(header[4], "十年期收益率", "E1")

    rows: list[tuple[dt.date, float, float]] = []
    for row_index, values in enumerate(rows_iter, start=2):
        if all(value is None or (isinstance(value, str) and not value.strip()) for value in values):
            continue

        date = _parse_date(values[0], epoch=workbook.epoch)
        try:
            yield_raw = _coerce_float(values[4])
        except ValueError as exc:
            raise ValueError(f"data_bond E{row_index} 内容错误：{exc}") from exc

        rows.append((date, yield_raw, _normalize_yield(yield_raw)))

    if not rows:
        raise ValueError("data_bond：没有可用数据行")

    rows.sort(key=lambda item: item[0])
    return rows


def _merge_by_bond_dates(
    bond_rows: list[tuple[dt.date, float, float]],
    pe_rows: list[tuple[dt.date, float, float]],
) -> list[tuple[dt.date, float, float, float]]:
    merged: list[tuple[dt.date, float, float, float]] = []
    pe_index = 0

    for bond_date, bond_yield_raw, bond_yield_decimal in bond_rows:
        while pe_index < len(pe_rows) and pe_rows[pe_index][0] < bond_date:
            pe_index += 1

        if pe_index >= len(pe_rows):
            raise ValueError("合并失败：data_PE 数据不足，无法继续对齐日期")

        pe_date, pe_value, pe_close = pe_rows[pe_index]
        if pe_date >= bond_date:
            merged.append((bond_date, bond_yield_raw, pe_value, pe_close))
            pe_index += 1
            continue

    if not merged:
        raise ValueError("合并失败：未生成任何对齐行")

    return merged


def _compute_erp_rows(
    merged_rows: list[tuple[dt.date, float, float, float]],
    bond_rows: list[tuple[dt.date, float, float]],
) -> list[list[object]]:
    bond_decimal_by_date = {date: decimal for date, _, decimal in bond_rows}
    output: list[list[object]] = [["日期", "十年国债收益率", "PE-TTM-S", "全A点位", "股权风险溢价"]]

    for date, yield_raw, pe_value, close_value in merged_rows:
        bond_yield_decimal = bond_decimal_by_date.get(date)
        if bond_yield_decimal is None:
            raise ValueError("内部错误：未找到收益率小数值")
        erp = (1.0 + 1.0 / pe_value) / (1.0 + bond_yield_decimal) - 1.0
        output.append([date.isoformat(), yield_raw, pe_value, close_value, erp])

    return output


def _write_csv(rows: list[list[object]], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as file_handle:
        writer = csv.writer(file_handle)
        for row in rows:
            writer.writerow([_cell_to_text(value) for value in row])


def _filter_columns(rows: list[list[object]], drop_names: set[str]) -> list[list[object]]:
    if not rows or not drop_names:
        return rows
    header = rows[0]
    keep_indices = [i for i, name in enumerate(header) if str(name) not in drop_names]
    if len(keep_indices) == len(header):
        return rows
    filtered: list[list[object]] = []
    for row in rows:
        filtered.append([row[i] for i in keep_indices if i < len(row)])
    return filtered


def _write_xlsx(rows: list[list[object]], path: Path, sheet_title: str) -> None:
    workbook_out = openpyxl.Workbook()
    sheet_out = workbook_out.active
    sheet_out.title = sheet_title
    sheet_out.freeze_panes = "B2"
    for row in rows:
        sheet_out.append([_round_for_output(value) for value in row])
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook_out.save(path)


def _process_ratio_file(source_path: Path, *, metric_header: str) -> list[list[object]]:
    workbook = openpyxl.load_workbook(source_path, data_only=True, read_only=True)
    sheet_name = workbook.sheetnames[0] if workbook.sheetnames else None
    if not sheet_name:
        raise ValueError(f"{source_path.name}：未找到可用工作表")
    sheet = workbook[sheet_name]

    last_col = 4  # A-D
    rows_iter = _iter_rows_values(sheet, last_col=last_col)
    header = next(rows_iter, None)
    if not header:
        raise ValueError(f"{source_path.name}：未找到标题行")

    header_a = _validate_header_cell(header[0])
    _ = _validate_header_cell(header[3])
    if "日期" not in header_a and header_a.lower() != "date":
        raise ValueError(f"{source_path.name}：A1 标题应为“日期”")

    rows: list[tuple[dt.date, float]] = []
    for row_index, values in enumerate(rows_iter, start=2):
        try:
            date = _parse_date(values[0], epoch=workbook.epoch)
            ratio = _coerce_float(values[3])
            rows.append((date, ratio))
        except Exception:
            continue

    if not rows:
        raise ValueError(f"{source_path.name}：清洗后没有可用数据行")

    rows.sort(key=lambda item: item[0])
    output: list[list[object]] = [[header_a, metric_header]]
    output.extend([[date.isoformat(), ratio] for date, ratio in rows])
    return output


def _process_us_index_file(source_path: Path, *, index_name: str) -> list[list[object]]:
    workbook = openpyxl.load_workbook(source_path, data_only=True, read_only=True)
    try:
        sheet_name = workbook.sheetnames[0] if workbook.sheetnames else None
        if not sheet_name:
            raise ValueError(f"{index_name}：未找到可用工作表")
        sheet = workbook[sheet_name]
        rows_iter = _iter_rows_values(sheet, last_col=8)  # A-H
        header = next(rows_iter, None)
        if not header:
            raise ValueError(f"{index_name}：未找到标题行")

        _ = _validate_header_cell(header[0])
        _ = _validate_header_cell(header[7])

        rows: list[tuple[dt.date, float]] = []
        for values in rows_iter:
            if all(value is None or (isinstance(value, str) and not value.strip()) for value in values):
                continue
            try:
                date = _parse_date(values[0], epoch=workbook.epoch)
                close = _coerce_float(values[7])
            except Exception:
                # 无效行直接删除（日期不合法、收盘点位空白或非数字）
                continue
            rows.append((date, close))

        if not rows:
            raise ValueError(f"{index_name}：清洗后没有可用数据行")

        rows.sort(key=lambda item: item[0])
        output: list[list[object]] = [["日期", "收盘点位"]]
        output.extend([[date.isoformat(), close] for date, close in rows])
        return output
    finally:
        workbook.close()


def _build_us_average_rows(
    clean_rows: list[list[object]],
    *,
    average_window: int,
    index_name: str,
) -> tuple[list[list[object]], int]:
    if not isinstance(average_window, int):
        raise ValueError("均线变量必须为整数")
    if average_window < 0 or average_window > 4000:
        raise ValueError("均线变量超出范围（0-4000）")

    effective_window = 1 if average_window == 0 else average_window
    data_rows = clean_rows[1:]
    if not data_rows:
        raise ValueError(f"{index_name}：没有可用数据行")

    dates: list[str] = []
    close_values: list[float] = []
    for row_index, row in enumerate(data_rows, start=2):
        try:
            date_text = str(row[0])
            _ = dt.date.fromisoformat(date_text)
            close_value = row[1]
            if not isinstance(close_value, (int, float)):
                raise ValueError("收盘点位不是数值")
            dates.append(date_text)
            close_values.append(float(close_value))
        except Exception as exc:
            raise ValueError(f"{index_name} 第 {row_index} 行数据不合法：{exc}") from exc

    ma_values = _moving_average(close_values, effective_window)
    output: list[list[object]] = [["日期", "收盘点位", "参考线"]]
    for index, date_text in enumerate(dates):
        ma_value = ma_values[index]
        if ma_value is None:
            continue
        output.append([date_text, close_values[index], ma_value])

    if len(output) <= 1:
        raise ValueError(
            f"{index_name} 数据不足：均线变量={average_window}（实际窗口={effective_window}）导致无可导出结果"
        )

    return output, effective_window


def _rolling_median(sorted_window: list[float]) -> float:
    size = len(sorted_window)
    if size == 0:
        raise ValueError("窗口为空")
    mid = size // 2
    if size % 2 == 1:
        return float(sorted_window[mid])
    return (float(sorted_window[mid - 1]) + float(sorted_window[mid])) / 2.0


def _rolling_stddevp(sum_values: float, sum_squares: float, size: int) -> float:
    if size <= 0:
        raise ValueError("窗口为空")
    mean = sum_values / size
    variance = (sum_squares / size) - (mean * mean)
    if variance < 0 and variance > -1e-12:
        variance = 0.0
    if variance < 0:
        raise ValueError("方差为负数（数值异常）")
    return math.sqrt(variance)


def _compute_erp_rolling_bands(
    erp_rows: list[list[object]],
    *,
    window_size: int = 2000,
    include_percentile: bool = False,
) -> list[list[object]]:
    if not isinstance(window_size, int) or window_size <= 0:
        raise ValueError("滚动窗口 n 必须为正整数")
    if not erp_rows or len(erp_rows) < 2:
        raise ValueError("ERP 数据为空")

    header = erp_rows[0]
    if len(header) < 5 or header[4] != "股权风险溢价":
        raise ValueError("ERP 表头不符合预期")

    data_rows = erp_rows[1:]
    if len(data_rows) < window_size:
        raise ValueError(f"数据不足：至少需要 {window_size} 行交易日数据")

    header = ["日期", "十年国债收益率", "PE-TTM-S", "全A点位", "股权风险溢价"]
    if include_percentile:
        header.append("股权风险溢价分位")
    header.extend(["+2σ", "+1σ", "中位数", "-1σ", "-2σ"])
    output: list[list[object]] = [header]

    sorted_window: list[float] = []
    queue: deque[float] = deque()
    sum_values = 0.0
    sum_squares = 0.0

    for index, row in enumerate(data_rows):
        erp_value = row[4]
        if not isinstance(erp_value, (int, float)):
            raise ValueError(f"ERP 第 {index + 2} 行数值类型不合法")
        erp_float = float(erp_value)

        insort(sorted_window, erp_float)
        queue.append(erp_float)
        sum_values += erp_float
        sum_squares += erp_float * erp_float

        if len(queue) > window_size:
            leaving = queue.popleft()
            sum_values -= leaving
            sum_squares -= leaving * leaving
            remove_index = bisect_left(sorted_window, leaving)
            if remove_index >= len(sorted_window) or sorted_window[remove_index] != leaving:
                raise ValueError("内部错误：滚动窗口移除失败")
            sorted_window.pop(remove_index)

        if len(queue) < window_size:
            continue

        median = _rolling_median(sorted_window)
        stddevp = _rolling_stddevp(sum_values, sum_squares, window_size)
        upper2 = median + 2 * stddevp
        upper1 = median + stddevp
        lower1 = median - stddevp
        lower2 = median - 2 * stddevp

        row_out: list[object] = [row[0], row[1], row[2], row[3], erp_float]
        if include_percentile:
            row_out.append(round(_rolling_percentile(sorted_window, erp_float), 1))
        row_out.extend([upper2, upper1, median, lower1, lower2])
        output.append(row_out)

    return output


def _compute_erp_interval_bands(
    erp_rows: list[list[object]],
    *,
    start_date: dt.date,
    end_date: dt.date | None,
) -> tuple[dt.date, dt.date, dt.date, dt.date, list[list[object]], float, float]:
    if not erp_rows or len(erp_rows) < 2:
        raise ValueError("ERP 数据为空")

    header = erp_rows[0]
    if len(header) < 5 or header[4] != "股权风险溢价":
        raise ValueError("ERP 表头不符合预期")

    data_rows = erp_rows[1:]
    dates: list[dt.date] = []
    for index, row in enumerate(data_rows, start=2):
        try:
            dates.append(dt.date.fromisoformat(str(row[0])))
        except ValueError as exc:
            raise ValueError(f"ERP 第 {index} 行日期无法解析") from exc

    if not dates:
        raise ValueError("ERP 数据为空")

    earliest = dates[0]
    latest = dates[-1]
    if end_date is None:
        end_date = latest
    if start_date < earliest:
        raise ValueError(f"起始日期过早：最早日期为 {earliest.isoformat()}")
    if start_date > latest:
        raise ValueError(f"起始日期过晚：最近日期为 {latest.isoformat()}")
    if end_date < earliest:
        raise ValueError(f"终止日期过早：最早日期为 {earliest.isoformat()}")
    if end_date > latest:
        raise ValueError(f"终止日期过晚：最近日期为 {latest.isoformat()}")

    start_index = bisect_left(dates, start_date)
    if start_index >= len(dates):
        raise ValueError(f"起始日期过晚：最近日期为 {latest.isoformat()}")

    end_index = bisect_right(dates, end_date) - 1
    if end_index < 0:
        raise ValueError(f"终止日期过早：最早日期为 {earliest.isoformat()}")

    actual_start = dates[start_index]
    actual_end = dates[end_index]
    if actual_start > actual_end:
        raise ValueError("起始日期不能晚于终止日期（自动调整后）")

    interval_rows = data_rows[start_index : end_index + 1]
    if not interval_rows:
        raise ValueError("区间内没有数据")

    erp_values: list[float] = []
    sum_values = 0.0
    sum_squares = 0.0
    for index, row in enumerate(interval_rows, start=start_index + 2):
        value = row[4]
        if not isinstance(value, (int, float)):
            raise ValueError(f"ERP 第 {index} 行数值类型不合法")
        value_float = float(value)
        erp_values.append(value_float)
        sum_values += value_float
        sum_squares += value_float * value_float

    sorted_values = sorted(erp_values)
    median = _rolling_median(sorted_values)
    stddevp = _rolling_stddevp(sum_values, sum_squares, len(erp_values))
    upper2 = median + 2 * stddevp
    upper1 = median + stddevp
    lower1 = median - stddevp
    lower2 = median - 2 * stddevp

    output: list[list[object]] = [
        [
            "日期",
            "十年国债收益率",
            "PE-TTM-S",
            "全A点位",
            "股权风险溢价",
            "股权风险溢价分位",
            "+2σ",
            "+1σ",
            "中位数",
            "-1σ",
            "-2σ",
        ]
    ]
    for row in interval_rows:
        erp_value = float(row[4])
        percentile = round(_rolling_percentile(sorted_values, erp_value), 1)
        output.append([row[0], row[1], row[2], row[3], erp_value, percentile, upper2, upper1, median, lower1, lower2])

    return earliest, latest, actual_start, actual_end, output, median, stddevp


@app.get("/")
def index() -> object:
    return app.send_static_file("index.html")


@app.get("/api/files")
def list_files() -> object:
    if not INPUT_DIR.exists():
        return jsonify({"files": []})

    files = sorted(
        p.name
        for p in INPUT_DIR.iterdir()
        if p.is_file() and p.suffix.lower() == ".xlsx"
        if not p.name.startswith("~$")
    )
    return jsonify({"files": files})


@app.get("/api/download/config")
def download_config() -> object:
    try:
        cfg = _download_load_config()
        return jsonify({"tasks": cfg.get("tasks", [])})
    except FileNotFoundError as exc:
        return jsonify({"error": str(exc)}), 404
    except Exception as exc:  # pragma: no cover
        return jsonify({"error": f"读取失败：{exc}"}), 500


@app.post("/api/download/run")
def download_run() -> object:
    data = request.get_json(force=True)
    task_ids = data.get("task_ids", [])
    urls = data.get("urls", {})
    trigger = data.get("trigger", "manual")

    try:
        cfg = _download_load_config()
        for task in cfg.get("tasks", []):
            if task["id"] in urls and urls[task["id"]]:
                task["url"] = urls[task["id"]]
        _download_save_config(cfg)

        started = download_runner.start(task_ids, urls, trigger)
        if not started:
            return jsonify({"ok": False, "message": "已有任务在运行"}), 409
        return jsonify({"ok": True})
    except FileNotFoundError as exc:
        return jsonify({"ok": False, "message": str(exc)}), 404
    except Exception as exc:  # pragma: no cover
        return jsonify({"ok": False, "message": f"启动失败：{exc}"}), 500


@app.get("/api/download/status")
def download_status() -> object:
    with DOWNLOAD_STATUS_LOCK:
        status_copy = json.loads(json.dumps(DOWNLOAD_STATUS))
    return jsonify(status_copy)


@app.post("/api/convert")
def convert_file() -> object:
    payload = request.get_json(silent=True) or {}
    filename = payload.get("filename")

    if not filename:
        return jsonify({"error": "缺少文件名"}), 400

    safe_name = Path(filename).name
    if safe_name != filename or not safe_name.lower().endswith(".xlsx") or safe_name.startswith("~$"):
        return jsonify({"error": "文件名不合法"}), 400

    source_path = INPUT_DIR / safe_name
    if not source_path.exists():
        return jsonify({"error": "文件不存在"}), 404

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    output_csv_path = OUTPUT_DIR / f"{source_path.stem}.csv"
    output_xlsx_path = OUTPUT_DIR / f"{source_path.stem}_processed.xlsx"

    try:
        process_xlsx_to_outputs(source_path, output_csv_path, output_xlsx_path)
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    except Exception as exc:  # pragma: no cover - surfaced to UI
        return jsonify({"error": f"转换失败：{exc}"}), 500

    return jsonify({"output_csv": output_csv_path.name, "output_xlsx": output_xlsx_path.name})


@app.post("/api/erp")
def generate_erp() -> object:
    payload = request.get_json(silent=True) or {}
    include_yield = bool(payload.get("include_yield", True))
    include_pe = bool(payload.get("include_pe", True))
    include_erp_percentile = bool(payload.get("include_erp_percentile", True))
    try:
        pe_path = _find_input_xlsx("data_PE")
        bond_path = _find_input_xlsx("data_bond")

        pe_rows = _process_data_pe(pe_path)
        bond_rows = _process_data_bond(bond_path)
        merged_rows = _merge_by_bond_dates(bond_rows, pe_rows)

        pe_clean_rows: list[list[object]] = [["日期", "PE-TTM-S", "全A点位"]] + [
            [date.isoformat(), pe, close] for date, pe, close in pe_rows
        ]
        bond_clean_rows: list[list[object]] = [["日期", "十年国债收益率"]] + [
            [date.isoformat(), yield_raw] for date, yield_raw, _ in bond_rows
        ]
        merged_clean_rows: list[list[object]] = [["日期", "十年国债收益率", "PE-TTM-S", "全A点位"]] + [
            [date.isoformat(), yield_raw, pe, close] for date, yield_raw, pe, close in merged_rows
        ]
        erp_rows = _compute_erp_rows(merged_rows, bond_rows)

        drop_names: set[str] = set()
        if not include_yield:
            drop_names.add("十年国债收益率")
        if not include_pe:
            drop_names.add("PE-TTM-S")
        if not include_erp_percentile:
            drop_names.add("股权风险溢价分位")

        pe_clean_rows = _filter_columns(pe_clean_rows, drop_names)
        bond_clean_rows = _filter_columns(bond_clean_rows, drop_names)
        merged_clean_rows = _filter_columns(merged_clean_rows, drop_names)
        erp_rows = _filter_columns(erp_rows, drop_names)

        output = {
            "data_PE_clean": "data_PE_clean.csv",
            "data_bond_clean": "data_bond_clean.csv",
            "merged": "merged.csv",
            "erp": "ERP.csv",
        }

        _write_csv(pe_clean_rows, OUTPUT_DIR / output["data_PE_clean"])
        _write_csv(bond_clean_rows, OUTPUT_DIR / output["data_bond_clean"])
        _write_csv(merged_clean_rows, OUTPUT_DIR / output["merged"])
        _write_csv(erp_rows, OUTPUT_DIR / output["erp"])

        return jsonify(
            {
                "outputs": {
                    "data_PE_clean_csv": output["data_PE_clean"],
                    "data_bond_clean_csv": output["data_bond_clean"],
                    "merged_csv": output["merged"],
                    "erp_csv": output["erp"],
                }
            }
        )
    except FileNotFoundError as exc:
        return jsonify({"error": str(exc)}), 404
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    except Exception as exc:  # pragma: no cover
        return jsonify({"error": f"生成失败：{exc}"}), 500

@app.post("/api/erp10y")
def generate_erp_10year() -> object:
    try:
        pe_path = _find_input_xlsx("data_PE")
        bond_path = _find_input_xlsx("data_bond")

        pe_rows = _process_data_pe(pe_path)
        bond_rows = _process_data_bond(bond_path)
        merged_rows = _merge_by_bond_dates(bond_rows, pe_rows)
        erp_rows = _compute_erp_rows(merged_rows, bond_rows)

        bands_rows = _compute_erp_rolling_bands(erp_rows, window_size=2000)

        csv_name = "ERP_10Year.csv"
        _write_csv(bands_rows, OUTPUT_DIR / csv_name)

        return jsonify({"output_csv": csv_name})
    except FileNotFoundError as exc:
        return jsonify({"error": str(exc)}), 404
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    except Exception as exc:  # pragma: no cover
        return jsonify({"error": f"生成失败：{exc}"}), 500


@app.post("/api/erprolling")
def generate_erp_rolling() -> object:
    payload = request.get_json(silent=True) or {}
    n = payload.get("n")
    include_yield = bool(payload.get("include_yield", True))
    include_pe = bool(payload.get("include_pe", True))
    include_erp_percentile = bool(payload.get("include_erp_percentile", True))

    try:
        if isinstance(n, str):
            try:
                n = int(n.strip())
            except ValueError as exc:
                raise ValueError("n 必须为整数") from exc
        if not isinstance(n, int):
            raise ValueError("n 必须为整数")
        if n < 1 or n > 4000:
            raise ValueError("n 超出范围（1-4000）")

        pe_path = _find_input_xlsx("data_PE")
        bond_path = _find_input_xlsx("data_bond")

        pe_rows = _process_data_pe(pe_path)
        bond_rows = _process_data_bond(bond_path)
        merged_rows = _merge_by_bond_dates(bond_rows, pe_rows)
        erp_rows = _compute_erp_rows(merged_rows, bond_rows)

        bands_rows = _compute_erp_rolling_bands(erp_rows, window_size=n, include_percentile=include_erp_percentile)

        drop_names: set[str] = set()
        if not include_yield:
            drop_names.add("十年国债收益率")
        if not include_pe:
            drop_names.add("PE-TTM-S")
        if not include_erp_percentile:
            drop_names.add("股权风险溢价分位")
        bands_rows = _filter_columns(bands_rows, drop_names)

        csv_name = "ERP_Rolling Calculation.csv"
        _write_csv(bands_rows, OUTPUT_DIR / csv_name)

        return jsonify({"output_csv": csv_name, "n": n})
    except FileNotFoundError as exc:
        return jsonify({"error": str(exc)}), 404
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    except Exception as exc:  # pragma: no cover
        return jsonify({"error": f"生成失败：{exc}"}), 500


@app.post("/api/erpinterval")
def generate_erp_interval() -> object:
    payload = request.get_json(silent=True) or {}
    start_date_raw = payload.get("start_date")
    end_date_raw = payload.get("end_date")
    include_yield = bool(payload.get("include_yield", True))
    include_pe = bool(payload.get("include_pe", True))
    include_erp_percentile = bool(payload.get("include_erp_percentile", True))

    try:
        if not isinstance(start_date_raw, str) or not start_date_raw.strip():
            raise ValueError("缺少起始日期 start_date")
        try:
            start_date = dt.date.fromisoformat(start_date_raw.strip())
        except ValueError as exc:
            raise ValueError("起始日期格式必须为 YYYY-MM-DD") from exc

        end_date_defaulted = False
        if end_date_raw is None or (isinstance(end_date_raw, str) and not end_date_raw.strip()):
            end_date = None
            end_date_defaulted = True
        else:
            if not isinstance(end_date_raw, str):
                raise ValueError("终止日期格式必须为 YYYY-MM-DD")
            try:
                end_date = dt.date.fromisoformat(end_date_raw.strip())
            except ValueError as exc:
                raise ValueError("终止日期格式必须为 YYYY-MM-DD") from exc

        pe_path = _find_input_xlsx("data_PE")
        bond_path = _find_input_xlsx("data_bond")

        pe_rows = _process_data_pe(pe_path)
        bond_rows = _process_data_bond(bond_path)
        merged_rows = _merge_by_bond_dates(bond_rows, pe_rows)
        erp_rows = _compute_erp_rows(merged_rows, bond_rows)

        earliest, latest, actual_start, actual_end, output_rows, median, stddevp = _compute_erp_interval_bands(
            erp_rows, start_date=start_date, end_date=end_date
        )

        drop_names: set[str] = set()
        if not include_yield:
            drop_names.add("十年国债收益率")
        if not include_pe:
            drop_names.add("PE-TTM-S")
        if not include_erp_percentile:
            drop_names.add("股权风险溢价分位")
        output_rows = _filter_columns(output_rows, drop_names)

        csv_name = "ERP_Interval.csv"
        csv_path = OUTPUT_DIR / csv_name
        _write_csv(output_rows, csv_path)
        remote_url = _publish_csv_to_cos(csv_path, csv_name)

        adjusted = actual_start != start_date
        adjusted_end = (end_date is not None) and (actual_end != end_date)
        return jsonify(
            {
                "output_csv": csv_name,
                "input_start_date": start_date.isoformat(),
                "used_start_date": actual_start.isoformat(),
                "input_end_date": "" if end_date_defaulted else end_date.isoformat(),
                "used_end_date": actual_end.isoformat(),
                "earliest_date": earliest.isoformat(),
                "latest_date": latest.isoformat(),
                "adjusted_to_trading_day": adjusted,
                "adjusted_end_to_trading_day": adjusted_end,
                "end_date_defaulted": end_date_defaulted,
                "median": median,
                "stddevp": stddevp,
                "remote_url": remote_url,
            }
        )
    except FileNotFoundError as exc:
        return jsonify({"error": str(exc)}), 404
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    except Exception as exc:  # pragma: no cover
        return jsonify({"error": f"生成失败：{exc}"}), 500


@app.post("/api/thermometer/clean")
def generate_thermometer_clean() -> object:
    try:
        gdp_path = _find_input_xlsx("data_Ratio GDP")
        volume_path = _find_input_xlsx("data_Ratio Volume")
        lend_path = _find_input_xlsx("data_Ratio Securities Lend")

        gdp_rows = _process_ratio_file(gdp_path, metric_header="总市值/GDP")
        volume_rows = _process_ratio_file(volume_path, metric_header="成交量/总市值")
        lend_rows = _process_ratio_file(lend_path, metric_header="融资融券/总市值")

        outputs = {
            "ratio_gdp": "Ratio_GDP.csv",
            "ratio_volume": "Ratio_Volume.csv",
            "ratio_securities_lend": "Ratio_Securities_Lend.csv",
        }

        _write_csv(gdp_rows, OUTPUT_DIR / outputs["ratio_gdp"])
        _write_csv(volume_rows, OUTPUT_DIR / outputs["ratio_volume"])
        _write_csv(lend_rows, OUTPUT_DIR / outputs["ratio_securities_lend"])

        return jsonify(
            {
                "outputs": {
                    "ratio_gdp_csv": outputs["ratio_gdp"],
                    "ratio_volume_csv": outputs["ratio_volume"],
                    "ratio_securities_lend_csv": outputs["ratio_securities_lend"],
                }
            }
        )
    except FileNotFoundError as exc:
        return jsonify({"error": str(exc)}), 404
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    except Exception as exc:  # pragma: no cover
        return jsonify({"error": f"生成失败：{exc}"}), 500


def _parse_average_window_payload(payload: dict, *, field_name: str = "average_window") -> int:
    raw = payload.get(field_name, 850)
    if isinstance(raw, str):
        text = raw.strip()
        if not text:
            raw = 850
        else:
            try:
                raw = int(text)
            except ValueError as exc:
                raise ValueError("均线变量必须为整数（0-4000）") from exc
    if not isinstance(raw, int):
        raise ValueError("均线变量必须为整数（0-4000）")
    if raw < 0 or raw > 4000:
        raise ValueError("均线变量超出范围（0-4000）")
    return raw


@app.post("/api/us/sp500/average")
def generate_sp500_average() -> object:
    payload = request.get_json(silent=True) or {}
    try:
        average_window = _parse_average_window_payload(payload)
        source_path = _find_input_xlsx("data_SP500")
        clean_rows = _process_us_index_file(source_path, index_name="标普500")
        output_rows, effective_window = _build_us_average_rows(
            clean_rows,
            average_window=average_window,
            index_name="标普500",
        )
        csv_name = "SP500_Average.csv"
        csv_path = OUTPUT_DIR / csv_name
        _write_csv(output_rows, csv_path)
        remote_url = _publish_csv_to_cos(csv_path, csv_name)
        return jsonify(
            {
                "output_csv": csv_name,
                "input_average_window": average_window,
                "effective_average_window": effective_window,
                "rows": len(output_rows) - 1,
                "remote_url": remote_url,
            }
        )
    except FileNotFoundError as exc:
        return jsonify({"error": str(exc)}), 404
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    except Exception as exc:  # pragma: no cover
        return jsonify({"error": f"生成失败：{exc}"}), 500


@app.post("/api/us/nasdaq/average")
def generate_nasdaq_average() -> object:
    payload = request.get_json(silent=True) or {}
    try:
        average_window = _parse_average_window_payload(payload)
        source_path = _find_input_xlsx("data_NASDAQ")
        clean_rows = _process_us_index_file(source_path, index_name="纳斯达克")
        output_rows, effective_window = _build_us_average_rows(
            clean_rows,
            average_window=average_window,
            index_name="纳斯达克",
        )
        csv_name = "NASDAQ_Average.csv"
        csv_path = OUTPUT_DIR / csv_name
        _write_csv(output_rows, csv_path)
        remote_url = _publish_csv_to_cos(csv_path, csv_name)
        return jsonify(
            {
                "output_csv": csv_name,
                "input_average_window": average_window,
                "effective_average_window": effective_window,
                "rows": len(output_rows) - 1,
                "remote_url": remote_url,
            }
        )
    except FileNotFoundError as exc:
        return jsonify({"error": str(exc)}), 404
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    except Exception as exc:  # pragma: no cover
        return jsonify({"error": f"生成失败：{exc}"}), 500


@app.post("/api/thermometer/percentiles")
def generate_thermometer_percentiles() -> object:
    payload = request.get_json(silent=True) or {}

    def get_int(name: str, *, min_value: int, max_value: int) -> int:
        raw = payload.get(name)
        if isinstance(raw, str):
            raw = raw.strip()
            if not raw:
                raise ValueError(f"缺少参数：{name}")
            try:
                raw = int(raw)
            except ValueError as exc:
                raise ValueError(f"{name} 必须为整数") from exc
        if not isinstance(raw, int):
            raise ValueError(f"{name} 必须为整数")
        if raw < min_value or raw > max_value:
            raise ValueError(f"{name} 超出范围（{min_value}-{max_value}）")
        return raw

    def get_mode(name: str) -> str:
        raw = payload.get(name, "auto")
        if isinstance(raw, str):
            text = raw.strip().lower()
            if text in ("auto", "custom"):
                return text
        raise ValueError(f"{name} 必须为 auto 或 custom")

    def get_bool(name: str, default: bool = False) -> bool:
        raw = payload.get(name)
        if raw is None:
            return default
        if isinstance(raw, bool):
            return raw
        if isinstance(raw, str):
            text = raw.strip().lower()
            if text in ("1", "true", "yes", "y", "on"):
                return True
            if text in ("0", "false", "no", "n", "off"):
                return False
        raise ValueError(f"{name} 必须为布尔值")

    try:
        ma_gdp = get_int("moving_average_gdp", min_value=1, max_value=1000)
        rp_gdp = get_int("rolling_period_gdp", min_value=1, max_value=1000)
        internal_gdp_mode = get_mode("internal_gdp_mode")
        internal_gdp = None if internal_gdp_mode == "auto" else get_int("internal_gdp", min_value=1, max_value=1000)
        ma_volume = get_int("moving_average_volume", min_value=1, max_value=4000)
        rp_volume = get_int("rolling_period_volume", min_value=1, max_value=4000)
        internal_volume_mode = get_mode("internal_volume_mode")
        internal_volume = None if internal_volume_mode == "auto" else get_int("internal_volume", min_value=1, max_value=4000)
        ma_securities = get_int("moving_average_securities", min_value=1, max_value=4000)
        rp_securities = get_int("rolling_period_securities", min_value=1, max_value=4000)
        internal_securities_mode = get_mode("internal_securities_mode")
        internal_securities = (
            None if internal_securities_mode == "auto" else get_int("internal_securities", min_value=1, max_value=4000)
        )
        ma_erp = get_int("moving_erp", min_value=1, max_value=4000)
        rp_erp = get_int("rolling_period_erp", min_value=1, max_value=4000)
        internal_erp_mode = get_mode("internal_erp_mode")
        internal_erp = None if internal_erp_mode == "auto" else get_int("internal_erp", min_value=1, max_value=4000)
        include_window_size = get_bool("include_window_size", False)

        if internal_gdp_mode == "custom" and internal_gdp is not None and internal_gdp > rp_gdp:
            raise ValueError("总市值/GDP最小递减滚动周期不能大于分位滚动周期")
        if internal_volume_mode == "custom" and internal_volume is not None and internal_volume > rp_volume:
            raise ValueError("成交量/总市值最小递减滚动周期不能大于分位滚动周期")
        if internal_securities_mode == "custom" and internal_securities is not None and internal_securities > rp_securities:
            raise ValueError("融资融券/总市值最小递减滚动周期不能大于分位滚动周期")
        if internal_erp_mode == "custom" and internal_erp is not None and internal_erp > rp_erp:
            raise ValueError("股权风险溢价最小递减滚动周期不能大于分位滚动周期")

        gdp_path = _find_input_xlsx("data_Ratio GDP")
        volume_path = _find_input_xlsx("data_Ratio Volume")
        lend_path = _find_input_xlsx("data_Ratio Securities Lend")

        gdp_dates, gdp_values = _load_ratio_series(gdp_path)
        vol_dates, vol_values = _load_ratio_series(volume_path)
        sec_dates, sec_values = _load_ratio_series(lend_path)
        erp_dates, erp_values, erp_yields, erp_pes, erp_closes = _load_erp_series()

        def build_output(
            dates: list[str],
            values: list[float],
            *,
            metric_header: str,
            ma_window: int,
            rp_window: int,
            internal_mode: str,
            min_window: int | None,
        ) -> list[list[object]]:
            ma_values = _moving_average(values, ma_window)
            if internal_mode == "auto":
                if include_window_size:
                    pct_values, window_sizes = _rolling_percentiles_with_min_window_sizes(
                        ma_values, rp_window, rp_window
                    )
                else:
                    pct_values = _rolling_percentiles(ma_values, rp_window)
                    window_sizes = None
            else:
                if min_window is None:
                    raise ValueError("最小递减滚动周期不能为空")
                if include_window_size:
                    pct_values, window_sizes = _rolling_percentiles_with_min_window_sizes(
                        ma_values, rp_window, min_window
                    )
                else:
                    pct_values = _rolling_percentiles_with_min_window(ma_values, rp_window, min_window)
                    window_sizes = None
            out: list[list[object]] = [["日期", metric_header, "平均移动", "分位"]]
            if include_window_size:
                out[0].append("滚动周期长度")
            for index, date_text in enumerate(dates):
                if pct_values[index] is None:
                    continue
                row = [date_text, values[index], ma_values[index], pct_values[index]]
                if include_window_size and window_sizes is not None:
                    row.append(window_sizes[index])
                out.append(row)
            return out

        gdp_out = build_output(
            gdp_dates,
            gdp_values,
            metric_header="总市值/GDP",
            ma_window=ma_gdp,
            rp_window=rp_gdp,
            internal_mode=internal_gdp_mode,
            min_window=internal_gdp,
        )
        gdp_out = _keep_weekly_latest_rows(gdp_out)
        vol_out = build_output(
            vol_dates,
            vol_values,
            metric_header="成交量/总市值",
            ma_window=ma_volume,
            rp_window=rp_volume,
            internal_mode=internal_volume_mode,
            min_window=internal_volume,
        )
        sec_out = build_output(
            sec_dates,
            sec_values,
            metric_header="融资融券/总市值",
            ma_window=ma_securities,
            rp_window=rp_securities,
            internal_mode=internal_securities_mode,
            min_window=internal_securities,
        )
        erp_ma_values = _moving_average(erp_values, ma_erp)
        if internal_erp_mode == "auto":
            if include_window_size:
                erp_pct_values, erp_window_sizes = _rolling_percentiles_with_min_window_sizes(
                    erp_ma_values, rp_erp, rp_erp
                )
            else:
                erp_pct_values = _rolling_percentiles(erp_ma_values, rp_erp)
                erp_window_sizes = None
        else:
            if internal_erp is None:
                raise ValueError("股权风险溢价最小递减滚动周期不能为空")
            if include_window_size:
                erp_pct_values, erp_window_sizes = _rolling_percentiles_with_min_window_sizes(
                    erp_ma_values, rp_erp, internal_erp
                )
            else:
                erp_pct_values = _rolling_percentiles_with_min_window(erp_ma_values, rp_erp, internal_erp)
                erp_window_sizes = None
        erp_out: list[list[object]] = [
            ["日期", "股权风险溢价", "平均移动", "分位", "十年国债收益率", "PE-TTM-S", "全A点位"]
        ]
        if include_window_size:
            erp_out[0].insert(4, "滚动周期长度")
        for index, date_text in enumerate(erp_dates):
            if erp_pct_values[index] is None:
                continue
            row = [
                date_text,
                erp_values[index],
                erp_ma_values[index],
                round(float(erp_pct_values[index]), 1),
                erp_yields[index],
                erp_pes[index],
                erp_closes[index],
            ]
            if include_window_size and erp_window_sizes is not None:
                row.insert(4, erp_window_sizes[index])
            erp_out.append(row)

        outputs = {
            "ratio_gdp": "Ratio_GDP_Percentile.csv",
            "ratio_volume": "Ratio_Volume_Percentile.csv",
            "ratio_securities_lend": "Ratio_Securities_Lend_Percentile.csv",
            "erp": "ERP_Percentile.csv",
        }

        _write_csv(gdp_out, OUTPUT_DIR / outputs["ratio_gdp"])
        _write_csv(vol_out, OUTPUT_DIR / outputs["ratio_volume"])
        _write_csv(sec_out, OUTPUT_DIR / outputs["ratio_securities_lend"])
        _write_csv(erp_out, OUTPUT_DIR / outputs["erp"])

        return jsonify(
            {
                "outputs": {
                    "ratio_gdp_csv": outputs["ratio_gdp"],
                    "ratio_volume_csv": outputs["ratio_volume"],
                    "ratio_securities_lend_csv": outputs["ratio_securities_lend"],
                    "erp_csv": outputs["erp"],
                }
            }
        )
    except FileNotFoundError as exc:
        return jsonify({"error": str(exc)}), 404
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    except Exception as exc:  # pragma: no cover
        return jsonify({"error": f"生成失败：{exc}"}), 500


@app.post("/api/thermometer/merge")
def generate_thermometer_merge() -> object:
    payload = request.get_json(silent=True) or {}

    def get_int(name: str, *, min_value: int, max_value: int) -> int:
        raw = payload.get(name)
        if isinstance(raw, str):
            raw = raw.strip()
            if not raw:
                raise ValueError(f"缺少参数：{name}")
            try:
                raw = int(raw)
            except ValueError as exc:
                raise ValueError(f"{name} 必须为整数") from exc
        if not isinstance(raw, int):
            raise ValueError(f"{name} 必须为整数")
        if raw < min_value or raw > max_value:
            raise ValueError(f"{name} 超出范围（{min_value}-{max_value}）")
        return raw

    def get_weight(name: str) -> float:
        raw = payload.get(name)
        if isinstance(raw, str):
            raw = raw.strip()
        try:
            value = float(raw)
        except Exception as exc:
            raise ValueError(f"{name} 必须为数值") from exc
        if value < 0 or value > 100:
            raise ValueError(f"{name} 超出范围（0-100）")
        return value

    def get_bool(name: str, default: bool) -> bool:
        raw = payload.get(name)
        if raw is None:
            return default
        if isinstance(raw, bool):
            return raw
        if isinstance(raw, str):
            text = raw.strip().lower()
            if text in ("1", "true", "yes", "y", "on"):
                return True
            if text in ("0", "false", "no", "n", "off"):
                return False
        raise ValueError(f"{name} 必须为布尔值")

    def get_mode(name: str) -> str:
        raw = payload.get(name, "auto")
        if isinstance(raw, str):
            text = raw.strip().lower()
            if text in ("auto", "custom"):
                return text
        raise ValueError(f"{name} 必须为 auto 或 custom")

    try:
        ma_gdp = get_int("moving_average_gdp", min_value=1, max_value=1000)
        rp_gdp = get_int("rolling_period_gdp", min_value=1, max_value=1000)
        internal_gdp_mode = get_mode("internal_gdp_mode")
        internal_gdp = None if internal_gdp_mode == "auto" else get_int("internal_gdp", min_value=1, max_value=1000)
        ma_volume = get_int("moving_average_volume", min_value=1, max_value=4000)
        rp_volume = get_int("rolling_period_volume", min_value=1, max_value=4000)
        internal_volume_mode = get_mode("internal_volume_mode")
        internal_volume = None if internal_volume_mode == "auto" else get_int("internal_volume", min_value=1, max_value=4000)
        ma_securities = get_int("moving_average_securities", min_value=1, max_value=4000)
        rp_securities = get_int("rolling_period_securities", min_value=1, max_value=4000)
        internal_securities_mode = get_mode("internal_securities_mode")
        internal_securities = (
            None if internal_securities_mode == "auto" else get_int("internal_securities", min_value=1, max_value=4000)
        )
        ma_erp = get_int("moving_erp", min_value=1, max_value=4000)
        rp_erp = get_int("rolling_period_erp", min_value=1, max_value=4000)
        internal_erp_mode = get_mode("internal_erp_mode")
        internal_erp = None if internal_erp_mode == "auto" else get_int("internal_erp", min_value=1, max_value=4000)

        weight_gdp = get_weight("weight_gdp")
        weight_volume = get_weight("weight_volume")
        weight_securities = get_weight("weight_securities_lend")
        weight_erp = get_weight("weight_erp")
        weight_sum = weight_gdp + weight_volume + weight_securities + weight_erp
        if weight_sum > 100.0 + 1e-9:
            raise ValueError("权重之和不能超过 100%")

        include_gdp = get_bool("include_gdp_percentile", True)
        include_volume = get_bool("include_volume_percentile", True)
        include_securities = get_bool("include_securities_percentile", True)
        include_erp = get_bool("include_erp", True)
        include_yield = get_bool("include_bond_yield", True)

        gdp_path = _find_input_xlsx("data_Ratio GDP")
        volume_path = _find_input_xlsx("data_Ratio Volume")
        lend_path = _find_input_xlsx("data_Ratio Securities Lend")

        gdp_dates, gdp_values = _load_ratio_series(gdp_path)
        vol_dates, vol_values = _load_ratio_series(volume_path)
        sec_dates, sec_values = _load_ratio_series(lend_path)
        erp_dates, erp_values, erp_yields, _, erp_closes = _load_erp_series()

        gdp_records_full = _build_percentile_records(
            gdp_dates,
            gdp_values,
            ma_window=ma_gdp,
            rp_window=rp_gdp,
            internal_mode=internal_gdp_mode,
            min_window=internal_gdp,
        )
        vol_records = _build_percentile_records(
            vol_dates,
            vol_values,
            ma_window=ma_volume,
            rp_window=rp_volume,
            internal_mode=internal_volume_mode,
            min_window=internal_volume,
        )
        sec_records = _build_percentile_records(
            sec_dates,
            sec_values,
            ma_window=ma_securities,
            rp_window=rp_securities,
            internal_mode=internal_securities_mode,
            min_window=internal_securities,
        )
        erp_records = _build_erp_percentile_records(
            erp_dates,
            erp_values,
            erp_yields,
            erp_closes,
            ma_window=ma_erp,
            rp_window=rp_erp,
            internal_mode=internal_erp_mode,
            min_window=internal_erp,
        )

        if not (gdp_records_full and vol_records and sec_records and erp_records):
            raise ValueError("数据不足：请检查移动平均与滚动周期参数是否过大")

        vol_start = vol_records[0][0]
        sec_start = sec_records[0][0]
        erp_start = erp_records[0]["date"]  # type: ignore[assignment]
        assert isinstance(erp_start, dt.date)

        vol_end = vol_records[-1][0]
        sec_end = sec_records[-1][0]
        erp_end = erp_records[-1]["date"]  # type: ignore[assignment]
        assert isinstance(erp_end, dt.date)
        gdp_end_full = gdp_records_full[-1][0]
        date_end = min(gdp_end_full, vol_end, sec_end, erp_end)

        gdp_records_filtered = [record for record in gdp_records_full if record[0] <= date_end]
        if not gdp_records_filtered:
            raise ValueError("合并失败：GDP 数据在有效区间内为空")

        gdp_records = _keep_weekly_latest_records(gdp_records_filtered)
        if not gdp_records:
            raise ValueError("合并失败：GDP 周频数据为空")

        date_begin = max(vol_start, sec_start, erp_start)
        gdp_dates_only = [d for d, _ in gdp_records]
        gdp_start_index = _nearest_index(gdp_dates_only, date_begin)
        start_date_used = gdp_dates_only[gdp_start_index]

        gdp_end_index = bisect_right(gdp_dates_only, date_end) - 1
        if gdp_end_index < gdp_start_index:
            raise ValueError("合并失败：有效时间区间为空")

        vol_dates_only = [d for d, _ in vol_records]
        sec_dates_only = [d for d, _ in sec_records]
        erp_dates_only = [record["date"] for record in erp_records]
        assert all(isinstance(d, dt.date) for d in erp_dates_only)
        erp_dates_only_typed: list[dt.date] = [d for d in erp_dates_only if isinstance(d, dt.date)]

        header = ["日期", "股权风险溢价分位", "全A点位", "市场温度"]
        if include_gdp:
            header.insert(1, "市值/GDP分位")
        if include_volume:
            header.insert(2 if include_gdp else 1, "成交量/市值分位")
        if include_securities:
            insert_at = 3 if include_gdp and include_volume else 2 if (include_gdp or include_volume) else 1
            header.insert(insert_at, "融资融券/市值分位")
        if include_erp:
            header.append("股权风险溢价")
        if include_yield:
            header.append("十年国债收益率")

        rows: list[list[object]] = [header]
        one_decimal_columns = {
            "市值/GDP分位",
            "成交量/市值分位",
            "融资融券/市值分位",
            "股权风险溢价分位",
            "市场温度",
            "全A点位",
        }

        def _get_percentile(records: list[tuple[dt.date, float]], dates_only: list[dt.date], target: dt.date) -> float:
            idx = _nearest_index(dates_only, target)
            return float(records[idx][1])

        for gdp_idx in range(gdp_start_index, gdp_end_index + 1):
            date_value = gdp_dates_only[gdp_idx]
            gdp_pct = float(gdp_records[gdp_idx][1])
            vol_pct = _get_percentile(vol_records, vol_dates_only, date_value)
            sec_pct = _get_percentile(sec_records, sec_dates_only, date_value)

            erp_idx = _nearest_index(erp_dates_only_typed, date_value)
            erp_record = erp_records[erp_idx]
            erp_pct = float(erp_record["erp_percentile"])
            close_value = float(erp_record["close"])
            erp_value = float(erp_record["erp"])
            yield_value = float(erp_record["yield"])

            temperature = (
                weight_gdp * gdp_pct
                + weight_volume * vol_pct
                + weight_securities * sec_pct
                + weight_erp * (100.0 - erp_pct)
            ) / 100.0

            row: dict[str, object] = {
                "日期": date_value.isoformat(),
                "市值/GDP分位": gdp_pct,
                "成交量/市值分位": vol_pct,
                "融资融券/市值分位": sec_pct,
                "股权风险溢价分位": erp_pct,
                "股权风险溢价": erp_value,
                "十年国债收益率": yield_value,
                "全A点位": close_value,
                "市场温度": temperature,
            }
            output_row: list[object] = []
            for col in header:
                value = row.get(col, "")
                if col in one_decimal_columns and isinstance(value, (int, float)) and not isinstance(value, bool):
                    value = round(float(value), 1)
                output_row.append(value)
            rows.append(output_row)

        output_name = "Market_Thermometer.csv"
        output_path = OUTPUT_DIR / output_name
        _write_csv(rows, output_path)
        remote_url = _publish_csv_to_cos(output_path, output_name)
        return jsonify(
            {
                "output_csv": output_name,
                "date_begin": date_begin.isoformat(),
                "date_begin_used": start_date_used.isoformat(),
                "date_end": date_end.isoformat(),
                "columns": header,
                "remote_url": remote_url,
            }
        )
    except FileNotFoundError as exc:
        return jsonify({"error": str(exc)}), 404
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    except Exception as exc:  # pragma: no cover
        return jsonify({"error": f"生成失败：{exc}"}), 500


if __name__ == "__main__":
    debug = os.environ.get("DP_DEBUG") == "1"
    app.run(host="127.0.0.1", port=5001, debug=debug, use_reloader=False)
